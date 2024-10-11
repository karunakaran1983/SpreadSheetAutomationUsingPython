import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Task 1: List each company with respective product count
# Expected Exercise Result: { 'AAA Company': 43, 'BBB Company': 17, 'CCC Company': 14 }

# print(product_list.max_row) # 75

products_per_supplier = {}
total_inventory_value_of_supplier = {}
inventory_less_than_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value
    total_inventory = product_list.cell(product_row, 5)

    if supplier_name in products_per_supplier:
        # current_num_products = products_per_supplier[supplier_name]
        current_num_products = products_per_supplier.get(supplier_name) # We could use dict.get() to get the value of the key
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        # print("Adding new supplier", supplier_name)
        products_per_supplier[supplier_name] = 1

    # Task 2: List each company with respective total inventory value
    if supplier_name in total_inventory_value_of_supplier:
        current_inventory_value = total_inventory_value_of_supplier.get(supplier_name)
        total_inventory_value_of_supplier[supplier_name] = current_inventory_value + inventory * price
    else:
        total_inventory_value_of_supplier[supplier_name] = inventory * price

    # Task 3: List of inventory which are less than 10
    if inventory < 10:
        inventory_less_than_10[int(product_number)] = int(inventory)

    # Task 4: Add inventory value (inventory * price)
    total_inventory.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")

print("Task 1: List each company with respective product count -> ", products_per_supplier)
print("Task 2: List each company with respective total inventory value -> ", total_inventory_value_of_supplier)
print("Task 3: List of inventory which are less than 10 -> ", inventory_less_than_10)
print("Task 4: Add inventory value (inventory * price) -> inventory_with_total_value.xlsx")